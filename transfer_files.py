import os
import shutil
from openpyxl import load_workbook as lwb

path_sourc = r"D:\python all\перенос файлов с именами из списка в отдельную папку\DIADOC"
path_finish = r"D:\python all\перенос файлов с именами из списка в отдельную папку\d"
file_exl = lwb(r"D:\python all\перенос файлов с именами из списка в отдельную папку\DIADOC.xlsx")
ws = file_exl.active

print(" Start reading directory")
df = os.listdir(path_sourc)
print(" Finish reading directory")

for i in range(2,ws.max_row + 1): # есть заголовок поэтому со второго
  print("Строка r = ", i)
  name_file = ws.cell(i, 1).value
  shutil.copy2((path_sourc+"\\"+name_file), (path_finish+"\\"+name_file) )
  